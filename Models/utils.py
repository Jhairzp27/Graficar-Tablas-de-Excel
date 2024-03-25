import openpyxl

def obtener_opciones_desde_excel(filepath, tipo_unidad):
    # Función para obtener las opciones de unidades desde un archivo Excel
    wb = openpyxl.load_workbook(filepath)
    hoja = wb["PlantillaVueltas"]
    opciones = []

    if tipo_unidad == "Grandes":
        # Obtener opciones para unidades Grandes desde la celda A5 hasta A30
        for row in range(5, 31):
            opcion = hoja.cell(row=row, column=1).value
            if opcion:
                opciones.append(opcion)
    elif tipo_unidad == "Micros":
        # Obtener opciones para Micros desde la celda A31 hasta A65
        for row in range(31, 66):
            opcion = hoja.cell(row=row, column=1).value
            if opcion:
                opciones.append(opcion)

    return opciones
