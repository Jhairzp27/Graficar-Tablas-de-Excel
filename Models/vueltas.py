import openpyxl
import matplotlib.pyplot as plt
import numpy as np

def generar_grafico_vueltas(filepath, tipo_unidad, num_unidad, ax=None):
    # Cargar el archivo Excel
    wb = openpyxl.load_workbook(filepath)
    hoja = wb["PlantillaVueltas"]

    # Definir variables con valores predeterminados
    rango_filas_lv = None
    rango_filas_fs = None
    columna_prog_lv = None
    columna_promedio_lv = None
    columna_ejec_lv = None
    columna_prog_fs = None
    columna_promedio_fs = None
    columna_ejec_fs = None

    # Determinar los rangos de filas y columnas para unidades Grandes y Micros
    if tipo_unidad == "Grandes":
        rango_filas_lv = (5, 30)
        rango_filas_fs = (5, 30)  # Mismo rango para unidades Grandes en fines de semana
        columna_prog_lv = 2
        columna_promedio_lv = 8
        columna_ejec_lv = 8
        columna_prog_fs = 3
        columna_promedio_fs = 9
        columna_ejec_fs = 9
        unidad_label = 'Número de Vueltas'
    elif tipo_unidad == "Micros":
        rango_filas_lv = (31, 65)
        rango_filas_fs = (31, 65)
        columna_prog_lv = 2
        columna_promedio_lv = 8
        columna_ejec_lv = 8
        columna_prog_fs = 3
        columna_promedio_fs = 9
        columna_ejec_fs = 9
        unidad_label = 'Número de Vueltas [Km]'
    else:
        print("Tipo de unidad no válido.")
        return

    # Filtrar los datos por número de unidad si se proporciona
    if num_unidad is not None:
        # Crear listas para almacenar los datos de las barras
        categorias = ['Programadas', 'Promedio', 'Ejecutadas']
        valores_lv = []
        valores_fs = []

        # Iterar sobre las filas para el gráfico de Lunes a Viernes
        for fila in range(rango_filas_lv[0], rango_filas_lv[1] + 1):
            unidad = hoja.cell(row=fila, column=1).value
            if unidad == num_unidad:
                valores_lv.append([
                    hoja.cell(row=fila, column=columna_prog_lv).value,
                    calcular_promedio(hoja, rango_filas_lv, columna_promedio_lv),
                    hoja.cell(row=fila, column=columna_ejec_lv).value
                ])

        # Iterar sobre las filas para el gráfico de Fines de Semana
        for fila in range(rango_filas_fs[0], rango_filas_fs[1] + 1):
            unidad = hoja.cell(row=fila, column=1).value
            if unidad == num_unidad:
                valores_fs.append([
                    hoja.cell(row=fila, column=columna_prog_fs).value,
                    calcular_promedio(hoja, rango_filas_fs, columna_promedio_fs),
                    hoja.cell(row=fila, column=columna_ejec_fs).value
                ])

        # Obtener los valores máximos para ambos gráficos
        max_lv = max(max(valores_lv))
        max_fs = max(max(valores_fs))

        # Ajustar el límite superior del eje y en función del valor máximo
        if max_lv > 4000:
            ylim_lv = max_lv + 10000
        elif max_lv < 2000:
            ylim_lv = max_lv + 200
        else:
            ylim_lv = max_lv

        if max_fs > 15000:  # Se cambia el límite superior para el gráfico de fines de semana
            ylim_fs = max_fs + 1995
        elif max_fs < 1000:
            ylim_fs = max_fs + 250  # Se cambia el valor sumado para el gráfico de fines de semana
        else:
            ylim_fs = max_fs

        # Calcular el límite superior final como el máximo de ambos
        ylim = max(ylim_lv, ylim_fs)

        # Crear el gráfico de barras en el eje proporcionado (o en uno nuevo si no se proporciona)
        if ax is None:
            (ax1, ax2) = plt.subplots(2, 1, figsize=(10, 8))
        else:
            ax1, ax2 = ax

        # Definir las ubicaciones de las barras en el eje x
        posiciones = np.arange(len(categorias))  # Posiciones para cada categoría
        bar_width = 0.28  # Anchura de las barras

        # Configurar los colores de las barras
        custom_colors = ['#87d349', '#FFA500', '#f9e826']
        # colores_lv = ['skyblue', 'orange', 'blue']
        # colores_fs = ['skyblue', 'orange', 'blue']

        # Gráfico 1: Lunes a Viernes
        for i, valores in enumerate(valores_lv):
            for j, valor in enumerate(valores):
                ax1.bar(posiciones[j] + i * bar_width, valor, width=bar_width, color=custom_colors[j])

                # Agregar etiquetas de valores a las barras
                ax1.text(posiciones[j] + i * bar_width, valor + 50, "{:.2f}".format(valor), ha='center', va='bottom')

        ax1.set_ylabel(unidad_label)
        ax1.set_title(f'Vueltas L-V - Unidad {num_unidad}')
        ax1.set_xticks(posiciones + bar_width * (len(valores_lv) - 1) / 2)  # Posiciones centradas
        ax1.set_xticklabels(categorias)
        ax1.set_ylim(0, ylim)  # Establecer límite superior del eje y

        # Gráfico 2: Fines de Semana
        for i, valores in enumerate(valores_fs):
            for j, valor in enumerate(valores):
                ax2.bar(posiciones[j] + i * bar_width, valor, width=bar_width, color=custom_colors[j])

                # Agregar etiquetas de valores a las barras
                ax2.text(posiciones[j] + i * bar_width, valor + 50, "{:.2f}".format(valor), ha='center', va='bottom')

        ax2.set_ylabel(unidad_label)
        ax2.set_title(f'Vueltas S-D - Unidad {num_unidad}')
        ax2.set_xticks(posiciones + bar_width * (len(valores_fs) - 1) / 2)  # Posiciones centradas
        ax2.set_xticklabels(categorias)
        ax2.set_ylim(0, ylim_fs)  # Establecer límite superior del eje y

        # Ajustar el diseño del gráfico para evitar solapamiento de etiquetas en el eje x
        plt.tight_layout()

        if ax is None:
            plt.show()
            plt.close()
        else:
            return ax1, ax2

    else:
        print("Número de unidad no proporcionado.")


def calcular_promedio(hoja, rango_filas, columna_promedio):
    suma_valores = 0
    num_valores = 0

    # Iterar sobre las filas y calcular el promedio
    for fila in range(rango_filas[0], rango_filas[1] + 1):
        celda = hoja.cell(row=fila, column=columna_promedio).value
        if isinstance(celda, (int, float)):
            suma_valores += celda
            num_valores += 1

    if num_valores > 0:
        promedio = suma_valores / num_valores
    else:
        promedio = 0

    return promedio
